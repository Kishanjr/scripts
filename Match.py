import pandas as pd
from openpyxl import load_workbook

# Load the Excel file
file_path = "your_excel_file.xlsx"
sheet_with_emails = "Sheet1"  # Replace with the sheet containing the list of emails

# Read the sheet with emails into a pandas DataFrame
emails_df = pd.read_excel(file_path, sheet_name=sheet_with_emails)

# Preprocess emails: Remove domain part and normalize
list_of_emails = emails_df['Email'].str.split('@').str[0].str.strip().str.lower().tolist()

# Load the workbook
wb = load_workbook(file_path)

# Function to preprocess and compare emails
def is_email_match(cell_value, email_list):
    # Remove the domain part from the email
    local_part = cell_value.split('@')[0].strip().lower()
    return local_part in email_list

# Process all sheets except the one containing the emails
for sheet_name in wb.sheetnames:
    if sheet_name == sheet_with_emails:
        continue  # Skip the sheet with the email list

    ws = wb[sheet_name]
    print(f"Processing sheet: {sheet_name}")
    
    # Create a new column header for the flag
    ws.cell(row=1, column=ws.max_column + 1, value="Flag")  # Add "Flag" column

    # Add flags for each email in Column D (4th column)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4), start=2):
        for cell in row:
            flag_cell = ws.cell(row=row_idx, column=ws.max_column)  # Add flag in the new column
            if cell.value and is_email_match(str(cell.value), list_of_emails):
                flag_cell.value = "Yes"  # Mark as matched
            else:
                flag_cell.value = "No"  # Mark as not matched

# Save the updated file
wb.save("updated_excel_file_with_flags.xlsx")
print("Flag column added and saved as 'updated_excel_file_with_flags.xlsx'.")
