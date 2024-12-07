import pandas as pd
from openpyxl import load_workbook

# File path
file_path = "output.xlsx"

# Function to simulate scraping (replace with your actual logic)
def scrape():
    # Simulated data fetched from a scraping process
    data = [
        {"Name": "Alice", "Age": 21, "City": "New York"},
        {"Name": "Bob", "Age": 23, "City": "Los Angeles"},
        {"Name": "Charlie", "Age": 24, "City": "Chicago"}
    ]
    return data

# Main function
def main():
    for iteration in range(3):  # Simulating 3 scraping iterations
        # Fetch data via the scrape function
        data = scrape()

        # Add dynamic `app` and `server` fields
        app = f"App_{iteration}"
        server = f"Server_{iteration}"
        for row in data:
            row["App"] = app
            row["Server"] = server

        # Convert JSON to DataFrame
        df = pd.DataFrame(data)

        # Define dynamic sheet name
        sheet_name = f"{app}"  # Use `app` as the dynamic sheet name

        try:
            # Open the workbook if it exists
            book = load_workbook(file_path)
            with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                # If sheet exists, find the next empty row
                if sheet_name in book.sheetnames:
                    sheet = book[sheet_name]
                    start_row = sheet.max_row
                    df.to_excel(writer, index=False, header=False, startrow=start_row, sheet_name=sheet_name)
                else:
                    # Create new sheet and write data
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
        except FileNotFoundError:
            # If the file does not exist, create it and write data
            with pd.ExcelWriter(file_path, engine="openpyxl", mode="w") as writer:
                df.to_excel(writer, index=False, sheet_name=sheet_name)

        print(f"Iteration {iteration}: Data written to sheet '{sheet_name}' in Excel.")

# Run the script
if __name__ == "__main__":
    main()
